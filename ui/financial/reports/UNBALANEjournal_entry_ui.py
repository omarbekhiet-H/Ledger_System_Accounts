# ui/financial/reports/journal_entry_detailed_report_ui.py

import sys
import os
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QPushButton, QLabel, QTableWidget, QTableWidgetItem,
    QHeaderView, QMessageBox, QDateEdit, QGroupBox, QGridLayout, QApplication,
    QAbstractItemView, QStyle, QHBoxLayout, QCheckBox
)
from PyQt5.QtCore import Qt, QDate
from PyQt5.QtGui import QFont, QColor

# =====================================================================
# تصحيح مسار المشروع الجذر لتمكين الاستيراد الصحيح
# =====================================================================
current_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.abspath(os.path.join(current_dir, '..', '..', '..'))
if project_root not in sys.path:
    sys.path.append(project_root)

# استيراد المدير الجديد وقاعدة البيانات
from database.manager.financial.transactions.journal_entry_detailed_manager import JournalEntryDetailedManager
from database.db_connection import get_financials_db_connection

# =====================================================================
# دالة مساعدة لتحميل ملف التصميم
# =====================================================================
def load_qss_file(file_path):
    """Loads content from a QSS file."""
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            return f.read()
    except FileNotFoundError:
        print(f"ERROR: QSS file not found at {file_path}")
        return ""
    except Exception as e:
        print(f"ERROR: Could not load QSS file {file_path}: {e}")
        return ""

class UnBlanceJournalEntryDetailedReportWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.manager = JournalEntryDetailedManager(get_financials_db_connection)
        self.setWindowTitle("تقرير قيود اليومية التفصيلي")
        self.setLayoutDirection(Qt.RightToLeft)
        self.init_ui()
        # ملاحظة: سيتم تطبيق التصميم من ملف التشغيل الرئيسي أو من قسم الاختبار أدناه

    def init_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(15, 15, 15, 15)
        main_layout.setSpacing(10)

        controls_group = QGroupBox("خيارات التقرير")
        controls_layout = QGridLayout(controls_group)
        
        controls_layout.addWidget(QLabel("من تاريخ:"), 0, 0)
        self.start_date_edit = QDateEdit(QDate.currentDate().addMonths(-1))
        self.start_date_edit.setCalendarPopup(True)
        self.start_date_edit.setDisplayFormat("yyyy-MM-dd")
        controls_layout.addWidget(self.start_date_edit, 0, 1)

        controls_layout.addWidget(QLabel("إلى تاريخ:"), 0, 2)
        self.end_date_edit = QDateEdit(QDate.currentDate())
        self.end_date_edit.setCalendarPopup(True)
        self.end_date_edit.setDisplayFormat("yyyy-MM-dd")
        controls_layout.addWidget(self.end_date_edit, 0, 3)

        # زر فلترة القيود غير المتزنة
        self.unbalanced_filter_cb = QCheckBox("عرض القيود غير المتزنة فقط")
        self.unbalanced_filter_cb.setStyleSheet("""
            QCheckBox {
                font-weight: bold;
                color: #d9534f;
                padding: 5px;
            }
            QCheckBox::indicator {
                width: 16px;
                height: 16px;
            }
        """)
        controls_layout.addWidget(self.unbalanced_filter_cb, 1, 0, 1, 2)

        # أزرار التحكم
        buttons_layout = QHBoxLayout()
        
        self.generate_btn = QPushButton("توليد التقرير")
        self.generate_btn.setIcon(self.style().standardIcon(QStyle.SP_DialogApplyButton))
        self.generate_btn.clicked.connect(self.generate_report)
        buttons_layout.addWidget(self.generate_btn)

        self.export_btn = QPushButton("تصدير إلى Excel")
        self.export_btn.setIcon(self.style().standardIcon(QStyle.SP_DialogSaveButton))
        self.export_btn.clicked.connect(self.export_to_excel)
        buttons_layout.addWidget(self.export_btn)

        controls_layout.addLayout(buttons_layout, 1, 2, 1, 2)
        
        main_layout.addWidget(controls_group)

        self.report_table = QTableWidget()
        self.report_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.report_table.setColumnCount(14)
        self.report_table.setHorizontalHeaderLabels([
            "رقم القيد", "التاريخ", "الوصف (القيد)", "نوع الحركة", "العملة", "الحساب",
            "مدين", "دائن", "رقم المستند (البند)", "نوع المستند (البند)",
            "مركز التكلفة", "نوع الضرائب", "ملاحظات (البند)", "الحالة"
        ])
        
        header = self.report_table.horizontalHeader()
        header.setSectionResizeMode(2, QHeaderView.Stretch)
        header.setSectionResizeMode(5, QHeaderView.Stretch)
        header.setSectionResizeMode(12, QHeaderView.Stretch)
        main_layout.addWidget(self.report_table)

    def generate_report(self):
        self.report_table.setRowCount(0)
        start_date = self.start_date_edit.date().toString("yyyy-MM-dd")
        end_date = self.end_date_edit.date().toString("yyyy-MM-dd")

        all_lines = self.manager.get_detailed_entries_in_range(start_date, end_date)

        if not all_lines:
            QMessageBox.information(self, "لا توجد بيانات", "لا توجد قيود يومية في الفترة المحددة.")
            return

        entries = {}
        for line in all_lines:
            entry_id = line['entry_id']
            if entry_id not in entries:
                entries[entry_id] = {'header': line, 'lines': []}
            entries[entry_id]['lines'].append(line)

        # فلترة القيود غير المتزنة إذا كان الخيار مفعلاً
        if self.unbalanced_filter_cb.isChecked():
            entries = self.filter_unbalanced_entries(entries)

        row = 0
        total_debit = 0.0
        total_credit = 0.0

        for entry_id, data in entries.items():
            header_data = data['header']
            
            self.report_table.insertRow(row)
            header_font = QFont(); header_font.setBold(True)
            header_color = QColor("#E0E6F0")

            self.report_table.setItem(row, 0, QTableWidgetItem(header_data['entry_number']))
            self.report_table.setItem(row, 1, QTableWidgetItem(header_data['entry_date']))
            self.report_table.setItem(row, 2, QTableWidgetItem(header_data['entry_description']))
            self.report_table.setItem(row, 3, QTableWidgetItem(header_data.get('transaction_type_name', '')))
            self.report_table.setItem(row, 4, QTableWidgetItem(header_data.get('currency_code', '')))
            self.report_table.setItem(row, 13, QTableWidgetItem(header_data.get('status', '')))

            for col in range(self.report_table.columnCount()):
                item = self.report_table.item(row, col)
                if not item: item = QTableWidgetItem("")
                item.setBackground(header_color)
                item.setFont(header_font)
                self.report_table.setItem(row, col, item)
            row += 1

            for line_data in data['lines']:
                self.report_table.insertRow(row)
                
                account_display = f"{line_data.get('acc_code', '')} - {line_data.get('account_name_ar', '')}"
                self.report_table.setItem(row, 5, QTableWidgetItem(account_display))
                
                debit = line_data.get('debit', 0.0)
                credit = line_data.get('credit', 0.0)
                self.add_numeric_item(row, 6, debit)
                self.add_numeric_item(row, 7, credit)
                
                self.report_table.setItem(row, 8, QTableWidgetItem(line_data.get('line_document_number') or ''))
                self.report_table.setItem(row, 9, QTableWidgetItem(line_data.get('document_type_name') or ''))
                self.report_table.setItem(row, 10, QTableWidgetItem(line_data.get('cost_center_name') or ''))
                self.report_table.setItem(row, 11, QTableWidgetItem(line_data.get('tax_type_name') or ''))
                self.report_table.setItem(row, 12, QTableWidgetItem(line_data.get('line_notes') or ''))
                
                total_debit += debit
                total_credit += credit
                row += 1
            
            self.report_table.insertRow(row)
            row += 1

        self.add_totals_row(row, total_debit, total_credit)
        self.report_table.resizeColumnsToContents()

    def filter_unbalanced_entries(self, entries):
        """فلترة القيود غير المتزنة فقط"""
        unbalanced_entries = {}
        
        for entry_id, data in entries.items():
            entry_debit = 0.0
            entry_credit = 0.0
            
            for line in data['lines']:
                entry_debit += line.get('debit', 0.0)
                entry_credit += line.get('credit', 0.0)
            
            # إذا كان القيد غير متزن (المدين ≠ الدائن)
            if abs(entry_debit - entry_credit) > 0.01:  # هامش خطأ صغير
                unbalanced_entries[entry_id] = data
        
        return unbalanced_entries

    def add_numeric_item(self, row, col, value):
        item = QTableWidgetItem(f"{value:,.2f}")
        item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.report_table.setItem(row, col, item)

    def add_totals_row(self, row, total_debit, total_credit):
        self.report_table.insertRow(row)
        total_font = QFont(); total_font.setBold(True)
        
        # تحديد لون الخلفية بناءً على التوازن
        if abs(total_debit - total_credit) > 0.01:
            total_color = QColor("#F8D7DA")  # أحمر فاتح للقيود غير المتزنة
        else:
            total_color = QColor("#D4EDDA")  # أخضر فاتح للقيود المتزنة
        
        self.report_table.setItem(row, 5, QTableWidgetItem("الإجمالي النهائي"))
        self.add_numeric_item(row, 6, total_debit)
        self.add_numeric_item(row, 7, total_credit)
        
        for col in range(self.report_table.columnCount()):
            item = self.report_table.item(row, col)
            if not item: item = QTableWidgetItem("")
            item.setBackground(total_color)
            item.setFont(total_font)
            self.report_table.setItem(row, col, item)

    def export_to_excel(self):
        """تصدير التقرير إلى Excel (وظيفة أساسية)"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = Workbook()
            ws = wb.active
            ws.title = "تقرير القيود التفصيلي"
            
            # إضافة العناوين
            headers = [
                "رقم القيد", "التاريخ", "الوصف (القيد)", "نوع الحركة", "العملة", "الحساب",
                "مدين", "دائن", "رقم المستند (البند)", "نوع المستند (البند)",
                "مركز التكلفة", "نوع الضرائب", "ملاحظات (البند)", "الحالة"
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color="E0E6F0", end_color="E0E6F0", fill_type="solid")
            
            # نسخ البيانات من الجدول
            for row in range(self.report_table.rowCount()):
                for col in range(self.report_table.columnCount()):
                    item = self.report_table.item(row, col)
                    if item:
                        ws.cell(row=row+2, column=col+1, value=item.text())
            
            # حفظ الملف
            import datetime
            filename = f"journal_detailed_report_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb.save(filename)
            
            QMessageBox.information(self, "تم التصدير", f"تم تصدير التقرير إلى ملف: {filename}")
            
        except ImportError:
            QMessageBox.warning(self, "خطأ", "لم يتم تثبيت مكتبة openpyxl. قم بتثبيتها باستخدام: pip install openpyxl")
        except Exception as e:
            QMessageBox.critical(self, "خطأ في التصدير", f"حدث خطأ أثناء التصدير: {str(e)}")

# =====================================================================
# الجزء الخاص بالتشغيل المستقل وتطبيق التصميم
# =====================================================================
if __name__ == '__main__':
    app = QApplication(sys.argv)

    # --- بناء مسار ملف التصميم بشكل صحيح ---
    # يبدأ من المسار الجذر للمشروع ثم يدخل إلى مجلدات التصميم
    qss_file_path = os.path.join(project_root, 'ui', 'styles', 'styles.qss')
    
    # --- تحميل وتطبيق ملف التصميم ---
    stylesheet = load_qss_file(qss_file_path)
    if stylesheet:
        app.setStyleSheet(stylesheet)
        print(f"DEBUG: Stylesheet loaded successfully from {qss_file_path}")
    else:
        print("WARNING: Failed to load stylesheet. UI will use default styles.")

    # --- إنشاء وعرض النافذة ---
    window = JournalEntryDetailedReportWindow()
    window.showMaximized()
    sys.exit(app.exec_())